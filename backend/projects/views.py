import logging
from rest_framework import generics, status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Count, Q
from django.utils import timezone

logger = logging.getLogger(__name__)

from .models import (
    Project, BOQ, BOQBill, BOQItem, Budget, BudgetRate, BudgetLineItem,
    IPC, IPCItem, ProjectRisk, ProjectVehicle, ProjectPersonnel, WeeklyProgress,
    ProjectPhase, ProjectActivity, ActivityProgress, VariationOrder,
)
from .serializers import (
    ProjectSerializer, ProjectDetailSerializer, BOQSerializer, BOQBillSerializer,
    BOQItemSerializer, BudgetSerializer, BudgetRateSerializer, BudgetLineItemSerializer,
    IPCSerializer, IPCItemSerializer, ProjectRiskSerializer, ProjectVehicleSerializer,
    ProjectPersonnelSerializer, WeeklyProgressSerializer,
    ProjectPhaseSerializer, ProjectActivitySerializer, ActivityProgressSerializer,
    VariationOrderSerializer,
)
from .services import BOQImportService, BudgetWorkbookImportService
import openpyxl
from decimal import Decimal, InvalidOperation
import datetime


def _safe_decimal(v):
    try:
        return Decimal(str(v)) if v not in (None, '') else Decimal('0')
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _safe_date(v):
    if not v:
        return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.date() if isinstance(v, datetime.datetime) else v
    try:
        return datetime.datetime.strptime(str(v).strip(), '%Y-%m-%d').date()
    except ValueError:
        try:
            return datetime.datetime.strptime(str(v).strip(), '%d/%m/%Y').date()
        except ValueError:
            return None


class ProjectImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Could not read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # Find header row — look for a row containing "code" or "project"
        header_row = None
        headers = {}
        for row in ws.iter_rows(max_row=10, values_only=True):
            row_lower = [str(c).strip().lower() if c else '' for c in row]
            if any(k in row_lower for k in ('code', 'project code', 'name', 'project name')):
                header_row = row_lower
                break

        if header_row is None:
            return Response({'error': 'Could not find header row. Expected columns: Code, Name, Client, Contract Number, Contract Value, Location, Start Date, End Date, Status'}, status=status.HTTP_400_BAD_REQUEST)

        # Map column names to indices
        col_map = {
            'code':             ['code', 'project code', 'proj code'],
            'name':             ['name', 'project name', 'project'],
            'client':           ['client', 'client name', 'employer'],
            'contract_number':  ['contract number', 'contract no', 'contract no.', 'contract#'],
            'contract_value':   ['contract value', 'value', 'contract amount', 'amount'],
            'location':         ['location', 'site', 'site location'],
            'start_date':       ['start date', 'start', 'commencement date'],
            'end_date':         ['end date', 'end', 'completion date', 'expected completion'],
            'status':           ['status'],
            'description':      ['description', 'notes', 'remarks'],
        }
        idx = {}
        for field, aliases in col_map.items():
            for i, h in enumerate(header_row):
                if h in aliases:
                    idx[field] = i
                    break

        if 'code' not in idx or 'name' not in idx:
            return Response({'error': f'Missing required columns. Found: {header_row}. Need at least "Code" and "Name".'}, status=status.HTTP_400_BAD_REQUEST)

        STATUS_MAP = {
            'planning': 'planning', 'plan': 'planning',
            'active': 'active', 'ongoing': 'active', 'in progress': 'active',
            'on hold': 'on_hold', 'hold': 'on_hold', 'on_hold': 'on_hold',
            'completed': 'completed', 'complete': 'completed', 'done': 'completed',
            'suspended': 'suspended', 'stopped': 'suspended',
        }

        created, updated, skipped = [], [], []
        data_started = False

        for row in ws.iter_rows(values_only=True):
            # Skip until past header
            row_vals = [str(c).strip().lower() if c else '' for c in row]
            if not data_started:
                if any(k in row_vals for k in ('code', 'project code', 'name')):
                    data_started = True
                continue

            def g(field, default=''):
                i = idx.get(field)
                if i is None or i >= len(row):
                    return default
                v = row[i]
                return str(v).strip() if v is not None else default

            code = g('code')
            name = g('name')
            if not code or not name:
                continue

            raw_status = g('status', 'planning').lower()
            proj_status = STATUS_MAP.get(raw_status, 'planning')

            defaults = {
                'name':            name,
                'client':          g('client'),
                'contract_number': g('contract_number'),
                'contract_value':  _safe_decimal(g('contract_value', '0')),
                'location':        g('location'),
                'start_date':      _safe_date(g('start_date')),
                'end_date':        _safe_date(g('end_date')),
                'status':          proj_status,
                'description':     g('description'),
            }

            try:
                obj, was_created = Project.objects.update_or_create(code=code, defaults=defaults)
                (created if was_created else updated).append(code)
            except Exception as e:
                skipped.append({'code': code, 'reason': str(e)})

        return Response({
            'created': len(created),
            'updated': len(updated),
            'skipped': len(skipped),
            'skipped_detail': skipped,
            'created_codes': created,
            'updated_codes': updated,
        }, status=status.HTTP_201_CREATED)


class BudgetWorkbookImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            result = BudgetWorkbookImportService.import_from_excel(file)
        except Exception as e:
            return Response({'error': f'Import failed: {e}'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result, status=status.HTTP_201_CREATED)


class ProjectListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer


class ProjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Project.objects.all()
    serializer_class = ProjectDetailSerializer
    lookup_field = 'pk'
    lookup_url_kwarg = 'project_pk'


class BOQListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BOQSerializer

    def get_queryset(self):
        return BOQ.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class BOQDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BOQSerializer

    def get_queryset(self):
        return BOQ.objects.filter(project_id=self.kwargs['project_pk'])


class BOQImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        title = request.data.get('title', file.name)
        result = BOQImportService.import_from_excel(file, project.id, title)
        return Response(result, status=status.HTTP_201_CREATED)


class BudgetListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetSerializer

    def get_queryset(self):
        return Budget.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class BudgetDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetSerializer

    def get_queryset(self):
        return Budget.objects.filter(project_id=self.kwargs['project_pk'])

    def get_object(self):
        return get_object_or_404(Budget, pk=self.kwargs['budget_pk'], project_id=self.kwargs['project_pk'])


class BudgetLineItemListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetLineItemSerializer

    def get_queryset(self):
        return BudgetLineItem.objects.filter(
            budget_id=self.kwargs['budget_pk'],
            budget__project_id=self.kwargs['project_pk'],
        )

    def perform_create(self, serializer):
        budget = get_object_or_404(Budget, pk=self.kwargs['budget_pk'], project_id=self.kwargs['project_pk'])
        serializer.save(budget=budget)


class BudgetLineItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetLineItemSerializer

    def get_queryset(self):
        return BudgetLineItem.objects.filter(
            budget_id=self.kwargs['budget_pk'],
            budget__project_id=self.kwargs['project_pk'],
        )


class BudgetSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_pk, budget_pk):
        budget = get_object_or_404(Budget, pk=budget_pk, project_id=project_pk)
        qs = BudgetLineItem.objects.filter(budget=budget)

        by_category = list(
            qs.values('category').annotate(
                base_total=Sum('base_cost'),
                high_total=Sum('high_case_cost'),
                low_total=Sum('low_case_cost'),
                count=Count('id')
            ).order_by('category')
        )

        # by_week aggregation across categories
        week_qs = qs.filter(week_no__isnull=False)
        by_week_raw = list(
            week_qs.values('week_no', 'category').annotate(
                base_total=Sum('base_cost'),
                high_total=Sum('high_case_cost'),
            ).order_by('week_no', 'category')
        )
        weeks = {}
        for row in by_week_raw:
            wn = row['week_no']
            if wn not in weeks:
                weeks[wn] = {'week_no': wn, 'materials': 0, 'fuel': 0, 'labour': 0, 'casuals': 0, 'base_total': 0, 'high_total': 0}
            cat = row['category']
            weeks[wn]['base_total'] = float(weeks[wn]['base_total']) + float(row['base_total'] or 0)
            weeks[wn]['high_total'] = float(weeks[wn]['high_total']) + float(row['high_total'] or 0)
            if cat in ('materials', 'fuel', 'labour', 'casuals'):
                weeks[wn][cat] = float(weeks[wn].get(cat, 0)) + float(row['base_total'] or 0)
        by_week = sorted(weeks.values(), key=lambda x: x['week_no'])

        by_month = list(
            qs.filter(month_no__isnull=False).values('month_no').annotate(
                base_total=Sum('base_cost'),
                high_total=Sum('high_case_cost'),
            ).order_by('month_no')
        )

        totals_agg = qs.aggregate(
            base=Sum('base_cost'),
            low=Sum('low_case_cost'),
            high=Sum('high_case_cost'),
            variance_reserve=Sum('variance_reserve'),
        )

        return Response({
            'by_category': by_category,
            'by_week': by_week,
            'by_month': by_month,
            'totals': {
                'base': float(totals_agg['base'] or 0),
                'low': float(totals_agg['low'] or 0),
                'high': float(totals_agg['high'] or 0),
                'variance_reserve': float(totals_agg['variance_reserve'] or 0),
            }
        })


class IPCListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = IPCSerializer

    def get_queryset(self):
        return IPC.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class IPCDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = IPCSerializer

    def get_queryset(self):
        return IPC.objects.filter(project_id=self.kwargs['project_pk'])

    def get_object(self):
        return get_object_or_404(IPC, pk=self.kwargs['ipc_pk'], project_id=self.kwargs['project_pk'])


class IPCItemListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = IPCItemSerializer

    def get_queryset(self):
        return IPCItem.objects.filter(
            ipc_id=self.kwargs['ipc_pk'],
            ipc__project_id=self.kwargs['project_pk'],
        )

    def perform_create(self, serializer):
        ipc = get_object_or_404(IPC, pk=self.kwargs['ipc_pk'], project_id=self.kwargs['project_pk'])
        serializer.save(ipc=ipc)


class IPCItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = IPCItemSerializer

    def get_queryset(self):
        return IPCItem.objects.filter(
            ipc_id=self.kwargs['ipc_pk'],
            ipc__project_id=self.kwargs['project_pk'],
        )


class ProjectRiskListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectRiskSerializer

    def get_queryset(self):
        return ProjectRisk.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class ProjectRiskDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectRiskSerializer

    def get_queryset(self):
        return ProjectRisk.objects.filter(project_id=self.kwargs['project_pk'])


class ProjectVehicleListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectVehicleSerializer

    def get_queryset(self):
        return ProjectVehicle.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        try:
            serializer.save(project=project)
        except Exception as e:
            logger.error(f"ProjectVehicle creation failed: {e}")
            raise


class ProjectVehicleDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectVehicleSerializer

    def get_queryset(self):
        return ProjectVehicle.objects.filter(project_id=self.kwargs['project_pk'])


class ProjectPersonnelListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectPersonnelSerializer

    def get_queryset(self):
        return ProjectPersonnel.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class ProjectPersonnelDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectPersonnelSerializer

    def get_queryset(self):
        return ProjectPersonnel.objects.filter(project_id=self.kwargs['project_pk'])


class WeeklyProgressListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WeeklyProgressSerializer

    def get_queryset(self):
        return WeeklyProgress.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class WeeklyProgressDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WeeklyProgressSerializer

    def get_queryset(self):
        return WeeklyProgress.objects.filter(project_id=self.kwargs['project_pk'])


class ProjectDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)

        # Budget summary
        budget_agg = BudgetLineItem.objects.filter(budget__project=project).aggregate(
            base_total=Sum('base_cost'),
            high_total=Sum('high_case_cost'),
            materials=Sum('base_cost', filter=Q(category='materials')),
            fuel=Sum('base_cost', filter=Q(category='fuel')),
            labour=Sum('base_cost', filter=Q(category='labour')),
            casuals=Sum('base_cost', filter=Q(category='casuals')),
        )

        # IPC summary
        ipc_agg = IPC.objects.filter(project=project).aggregate(
            total_claimed=Sum('amount_claimed'),
            total_certified=Sum('amount_certified'),
            total_paid=Sum('amount_paid'),
            ipc_count=Count('id'),
            pending_count=Count('id', filter=Q(status__in=['submitted', 'draft'])),
        )

        # Risk summary
        risk_agg = ProjectRisk.objects.filter(project=project).aggregate(
            open_count=Count('id', filter=Q(status='open')),
            high_count=Count('id', filter=Q(impact_level='high')),
            critical_count=Count('id', filter=Q(impact_level='critical')),
            mitigated_count=Count('id', filter=Q(status='mitigated')),
        )

        vehicle_count = ProjectVehicle.objects.filter(project=project).count()
        active_vehicle_count = ProjectVehicle.objects.filter(project=project, is_active=True).count()
        personnel_count = ProjectPersonnel.objects.filter(project=project).count()

        latest_progress = WeeklyProgress.objects.filter(project=project).order_by('-week_no').first()
        latest_progress_data = None
        if latest_progress:
            latest_progress_data = {
                'week_no': latest_progress.week_no,
                'total_actual': float(latest_progress.total_actual),
                'progress_notes': latest_progress.progress_notes,
            }

        weeks_elapsed = WeeklyProgress.objects.filter(project=project).count()

        return Response({
            'project': {
                'id': str(project.id),
                'code': project.code,
                'name': project.name,
                'client': project.client,
                'contract_number': project.contract_number,
                'contract_value': float(project.contract_value),
                'location': project.location,
                'status': project.status,
                'start_date': project.start_date,
                'end_date': project.end_date,
                'description': project.description,
            },
            'budget_summary': {
                'base_total': float(budget_agg['base_total'] or 0),
                'high_total': float(budget_agg['high_total'] or 0),
                'materials': float(budget_agg['materials'] or 0),
                'fuel': float(budget_agg['fuel'] or 0),
                'labour': float(budget_agg['labour'] or 0),
                'casuals': float(budget_agg['casuals'] or 0),
            },
            'ipc_summary': {
                'total_claimed': float(ipc_agg['total_claimed'] or 0),
                'total_certified': float(ipc_agg['total_certified'] or 0),
                'total_paid': float(ipc_agg['total_paid'] or 0),
                'ipc_count': ipc_agg['ipc_count'],
                'pending_count': ipc_agg['pending_count'],
            },
            'risk_summary': {
                'open_count': risk_agg['open_count'],
                'high_count': risk_agg['high_count'],
                'critical_count': risk_agg['critical_count'],
                'mitigated_count': risk_agg['mitigated_count'],
            },
            'fleet_summary': {
                'vehicle_count': vehicle_count,
                'active_count': active_vehicle_count,
            },
            'personnel_count': personnel_count,
            'latest_progress': latest_progress_data,
            'weeks_elapsed': weeks_elapsed,
        })


class ProjectCostingView(APIView):
    """Budget cost summary grouped by category for a project."""
    permission_classes = [IsAuthenticated]

    def get(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)

        qs = BudgetLineItem.objects.filter(budget__project=project)

        by_category = list(
            qs.values('category').annotate(
                base_total=Sum('base_cost'),
                high_total=Sum('high_case_cost'),
                low_total=Sum('low_case_cost'),
                variance_reserve=Sum('variance_reserve'),
                count=Count('id'),
            ).order_by('category')
        )

        totals_agg = qs.aggregate(
            base=Sum('base_cost'),
            low=Sum('low_case_cost'),
            high=Sum('high_case_cost'),
            variance_reserve=Sum('variance_reserve'),
        )

        # Actual spend from weekly progress
        progress_agg = WeeklyProgress.objects.filter(project=project).aggregate(
            materials_actual=Sum('materials_actual'),
            fuel_actual=Sum('fuel_actual'),
            labour_actual=Sum('labour_actual'),
            casuals_actual=Sum('casuals_actual'),
            total_actual=Sum('total_actual'),
        )

        return Response({
            'project_id': str(project.id),
            'project_code': project.code,
            'by_category': [
                {
                    'category': row['category'],
                    'count': row['count'],
                    'base_total': float(row['base_total'] or 0),
                    'high_total': float(row['high_total'] or 0),
                    'low_total': float(row['low_total'] or 0),
                    'variance_reserve': float(row['variance_reserve'] or 0),
                }
                for row in by_category
            ],
            'totals': {
                'base': float(totals_agg['base'] or 0),
                'low': float(totals_agg['low'] or 0),
                'high': float(totals_agg['high'] or 0),
                'variance_reserve': float(totals_agg['variance_reserve'] or 0),
            },
            'actual_spend': {
                'materials': float(progress_agg['materials_actual'] or 0),
                'fuel': float(progress_agg['fuel_actual'] or 0),
                'labour': float(progress_agg['labour_actual'] or 0),
                'casuals': float(progress_agg['casuals_actual'] or 0),
                'total': float(progress_agg['total_actual'] or 0),
            },
        })


# ---------------------------------------------------------------------------
# IPC Workflow Action Views
# ---------------------------------------------------------------------------

class IPCSubmitView(APIView):
    """POST /api/v1/projects/{project_pk}/ipcs/{ipc_pk}/submit/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, ipc_pk):
        ipc = get_object_or_404(IPC, pk=ipc_pk, project_id=project_pk)
        if ipc.status != 'draft':
            return Response(
                {'detail': f'Cannot submit IPC in status "{ipc.status}". Must be in draft.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ipc.status = 'submitted'
        ipc.submission_date = timezone.now().date()
        ipc.save(update_fields=['status', 'submission_date'])
        return Response({'detail': 'IPC submitted.', 'status': ipc.status})


class IPCCertifyView(APIView):
    """POST /api/v1/projects/{project_pk}/ipcs/{ipc_pk}/certify/
    Requires QS or Architect role (group membership).
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, ipc_pk):
        allowed_groups = {'QS', 'Architect', 'qs', 'architect', 'quantity_surveyor'}
        user_groups = set(request.user.groups.values_list('name', flat=True))
        if not (user_groups & allowed_groups) and not request.user.is_staff:
            return Response(
                {'detail': 'Only QS or Architect role can certify an IPC.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        ipc = get_object_or_404(IPC, pk=ipc_pk, project_id=project_pk)
        if ipc.status != 'submitted':
            return Response(
                {'detail': f'Cannot certify IPC in status "{ipc.status}". Must be submitted first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ipc.status = 'certified'
        ipc.certification_date = timezone.now().date()
        ipc.save(update_fields=['status', 'certification_date'])
        return Response({'detail': 'IPC certified.', 'status': ipc.status})


class IPCApproveView(APIView):
    """POST /api/v1/projects/{project_pk}/ipcs/{ipc_pk}/approve/
    Requires Finance role.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, ipc_pk):
        allowed_groups = {'Finance', 'finance'}
        user_groups = set(request.user.groups.values_list('name', flat=True))
        if not (user_groups & allowed_groups) and not request.user.is_staff:
            return Response(
                {'detail': 'Only Finance role can approve an IPC.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        ipc = get_object_or_404(IPC, pk=ipc_pk, project_id=project_pk)
        if ipc.status != 'certified':
            return Response(
                {'detail': f'Cannot approve IPC in status "{ipc.status}". Must be certified first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ipc.status = 'approved'
        ipc.save(update_fields=['status'])
        return Response({'detail': 'IPC approved.', 'status': ipc.status})


class IPCPayView(APIView):
    """POST /api/v1/projects/{project_pk}/ipcs/{ipc_pk}/pay/
    Requires Finance role.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, ipc_pk):
        allowed_groups = {'Finance', 'finance'}
        user_groups = set(request.user.groups.values_list('name', flat=True))
        if not (user_groups & allowed_groups) and not request.user.is_staff:
            return Response(
                {'detail': 'Only Finance role can mark an IPC as paid.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        ipc = get_object_or_404(IPC, pk=ipc_pk, project_id=project_pk)
        if ipc.status != 'approved':
            return Response(
                {'detail': f'Cannot mark IPC as paid in status "{ipc.status}". Must be approved first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ipc.status = 'paid'
        ipc.payment_date = timezone.now().date()
        ipc.save(update_fields=['status', 'payment_date'])
        return Response({'detail': 'IPC marked as paid.', 'status': ipc.status})


class IPCRejectView(APIView):
    """POST /api/v1/projects/{project_pk}/ipcs/{ipc_pk}/reject/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, ipc_pk):
        ipc = get_object_or_404(IPC, pk=ipc_pk, project_id=project_pk)
        if ipc.status not in ('submitted', 'certified', 'approved'):
            return Response(
                {'detail': f'Cannot reject IPC in status "{ipc.status}".'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {'detail': 'A rejection reason is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ipc.status = 'rejected'
        ipc.rejection_reason = reason
        ipc.save(update_fields=['status', 'rejection_reason'])
        return Response({'detail': 'IPC rejected.', 'status': ipc.status})


# ---------------------------------------------------------------------------
# Budget Approval Workflow Views
# ---------------------------------------------------------------------------

class BudgetSubmitView(APIView):
    """POST /api/v1/projects/{project_pk}/budgets/{budget_pk}/submit/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, budget_pk):
        budget = get_object_or_404(Budget, pk=budget_pk, project_id=project_pk)
        if budget.status not in ('draft', 'rejected'):
            return Response(
                {'detail': f'Cannot submit budget in status "{budget.status}". Must be draft or rejected.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        budget.status = 'pending_approval'
        budget.rejection_reason = ''
        budget.save(update_fields=['status', 'rejection_reason'])
        return Response({'detail': 'Budget submitted for approval.', 'status': budget.status})


class BudgetApproveView(APIView):
    """POST /api/v1/projects/{project_pk}/budgets/{budget_pk}/approve/
    Locks budget from further editing.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, budget_pk):
        if not request.user.is_staff:
            allowed_groups = {'Finance', 'finance', 'Project Manager', 'project_manager'}
            user_groups = set(request.user.groups.values_list('name', flat=True))
            if not (user_groups & allowed_groups):
                return Response(
                    {'detail': 'Only Finance or Project Manager role can approve a budget.'},
                    status=status.HTTP_403_FORBIDDEN,
                )
        budget = get_object_or_404(Budget, pk=budget_pk, project_id=project_pk)
        if budget.status != 'pending_approval':
            return Response(
                {'detail': f'Cannot approve budget in status "{budget.status}". Must be pending approval.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        budget.status = 'approved'
        budget.save(update_fields=['status'])
        return Response({'detail': 'Budget approved and locked from further edits.', 'status': budget.status})


class BudgetRejectView(APIView):
    """POST /api/v1/projects/{project_pk}/budgets/{budget_pk}/reject/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, budget_pk):
        if not request.user.is_staff:
            allowed_groups = {'Finance', 'finance', 'Project Manager', 'project_manager'}
            user_groups = set(request.user.groups.values_list('name', flat=True))
            if not (user_groups & allowed_groups):
                return Response(
                    {'detail': 'Only Finance or Project Manager role can reject a budget.'},
                    status=status.HTTP_403_FORBIDDEN,
                )
        budget = get_object_or_404(Budget, pk=budget_pk, project_id=project_pk)
        if budget.status != 'pending_approval':
            return Response(
                {'detail': f'Cannot reject budget in status "{budget.status}". Must be pending approval.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {'detail': 'A rejection reason is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        budget.status = 'rejected'
        budget.rejection_reason = reason
        budget.save(update_fields=['status', 'rejection_reason'])
        return Response({'detail': 'Budget rejected.', 'status': budget.status})


# ---------------------------------------------------------------------------
# Risk Register Status Update View
# ---------------------------------------------------------------------------

class ProjectRiskUpdateStatusView(APIView):
    """POST /api/v1/projects/{project_pk}/risks/{pk}/update-status/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk, pk):
        risk = get_object_or_404(ProjectRisk, pk=pk, project_id=project_pk)
        new_status = request.data.get('status', '').strip()
        valid_statuses = dict(ProjectRisk.STATUS_CHOICES).keys()
        if not new_status:
            return Response(
                {'detail': f'A status is required. Valid choices: {list(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if new_status not in valid_statuses:
            return Response(
                {'detail': f'Invalid status "{new_status}". Valid choices: {list(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        mitigation_notes = request.data.get('mitigation_notes', '').strip()
        risk.status = new_status
        if mitigation_notes:
            risk.notes = (risk.notes + '\n' + mitigation_notes).strip()
        risk.save(update_fields=['status', 'notes'])
        serializer = ProjectRiskSerializer(risk)
        return Response(serializer.data)


# ── WBS: Phases ──────────────────────────────────────────────────────────────

class ProjectPhaseListCreate(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectPhaseSerializer

    def get_queryset(self):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        return ProjectPhase.objects.filter(project=project).prefetch_related('activities')

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class ProjectPhaseDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectPhaseSerializer

    def get_queryset(self):
        return ProjectPhase.objects.filter(project_id=self.kwargs['project_pk'])


# ── WBS: Activities ───────────────────────────────────────────────────────────

class ProjectActivityListCreate(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectActivitySerializer

    def get_queryset(self):
        phase = get_object_or_404(ProjectPhase, pk=self.kwargs['phase_pk'], project_id=self.kwargs['project_pk'])
        return ProjectActivity.objects.filter(phase=phase)

    def perform_create(self, serializer):
        phase = get_object_or_404(ProjectPhase, pk=self.kwargs['phase_pk'], project_id=self.kwargs['project_pk'])
        serializer.save(phase=phase)


class ProjectActivityDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectActivitySerializer

    def get_queryset(self):
        return ProjectActivity.objects.filter(phase__project_id=self.kwargs['project_pk'])


# ── WBS: Activity Progress ────────────────────────────────────────────────────

class ActivityProgressListCreate(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ActivityProgressSerializer

    def get_queryset(self):
        activity = get_object_or_404(
            ProjectActivity, pk=self.kwargs['activity_pk'],
            phase__project_id=self.kwargs['project_pk']
        )
        return ActivityProgress.objects.filter(activity=activity)

    def perform_create(self, serializer):
        activity = get_object_or_404(
            ProjectActivity, pk=self.kwargs['activity_pk'],
            phase__project_id=self.kwargs['project_pk']
        )
        entry = serializer.save(activity=activity)
        # Update activity's percent_complete to match latest entry
        activity.percent_complete = entry.percent_complete
        if float(entry.percent_complete) >= 100 and activity.status != 'completed':
            activity.status = 'completed'
        elif float(entry.percent_complete) > 0 and activity.status == 'not_started':
            activity.status = 'in_progress'
        if float(entry.percent_complete) > 0 and not activity.actual_start:
            activity.actual_start = entry.date
        if float(entry.percent_complete) >= 100 and not activity.actual_end:
            activity.actual_end = entry.date
        activity.save()


class ProjectWBSSummaryView(APIView):
    """Returns overall project % complete derived from all phases/activities."""
    permission_classes = [IsAuthenticated]

    def get(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        phases = ProjectPhase.objects.filter(project=project).prefetch_related('activities')
        phase_data = ProjectPhaseSerializer(phases, many=True).data
        all_activities = ProjectActivity.objects.filter(phase__project=project)
        total_weight = sum(float(a.weight) for a in all_activities)
        if total_weight > 0:
            weighted = sum(float(a.weight) * float(a.percent_complete) for a in all_activities)
            overall_pct = round(weighted / total_weight, 1)
        else:
            overall_pct = 0
        return Response({
            'overall_percent_complete': overall_pct,
            'phase_count': phases.count(),
            'activity_count': all_activities.count(),
            'completed_activities': all_activities.filter(status='completed').count(),
            'in_progress_activities': all_activities.filter(status='in_progress').count(),
            'phases': phase_data,
        })


# ── WBS: Generate from BOQ ────────────────────────────────────────────────────
class GenerateWBSFromBOQView(APIView):
    """Auto-generate WBS phases and activities from the project's latest BOQ."""
    permission_classes = [IsAuthenticated]

    def post(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        boq = project.boqs.prefetch_related('bills__items').order_by('-uploaded_at').first()
        if not boq:
            return Response({'detail': 'No BOQ found for this project. Upload a BOQ first.'}, status=400)

        replace = request.data.get('replace', False)
        if replace:
            ProjectPhase.objects.filter(project=project).delete()

        COLORS = ['blue', 'green', 'purple', 'orange', 'teal', 'indigo', 'pink', 'red']
        created_phases = 0
        created_activities = 0

        for i, bill in enumerate(boq.bills.all()):
            phase_name = bill.description or f'Bill {bill.bill_number}'
            phase, created = ProjectPhase.objects.get_or_create(
                project=project,
                name=phase_name,
                defaults={
                    'description': f'Generated from BOQ Bill {bill.bill_number}',
                    'color': COLORS[i % len(COLORS)],
                    'order': i,
                    'planned_start': project.start_date,
                    'planned_end': project.end_date,
                }
            )
            if created:
                created_phases += 1

            for j, item in enumerate(bill.items.all()):
                wbs_code = f'{bill.bill_number}.{item.item_number}'
                act, act_created = ProjectActivity.objects.get_or_create(
                    phase=phase,
                    wbs_code=wbs_code,
                    defaults={
                        'description': item.description[:500],
                        'planned_start': project.start_date,
                        'planned_end': project.end_date,
                        'weight': float(item.amount) if float(item.amount or 0) > 0 else 1,
                        'status': 'not_started',
                        'order': j,
                    }
                )
                if act_created:
                    created_activities += 1

        return Response({
            'detail': f'Generated {created_phases} phases and {created_activities} activities from BOQ.',
            'created_phases': created_phases,
            'created_activities': created_activities,
        })


# ── Variation Orders ──────────────────────────────────────────────────────────
class VariationOrderListCreate(generics.ListCreateAPIView):
    serializer_class   = VariationOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return VariationOrder.objects.filter(project_id=self.kwargs['project_pk']).select_related('approved_by', 'created_by')

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project, created_by=self.request.user)


class VariationOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = VariationOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return VariationOrder.objects.filter(project_id=self.kwargs['project_pk'])


# ── EVM / Project Finance ─────────────────────────────────────────────────────
class EVMView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, project_pk):
        from django.db.models import Sum
        from datetime import date

        project = get_object_or_404(Project, pk=project_pk)
        today   = date.today()

        # ── Planned Value from approved budget ────────────────────────────────
        approved_budget = project.budgets.filter(status='approved').first()
        bac = 0.0
        pv  = 0.0
        by_category = {}

        if approved_budget:
            items = BudgetLineItem.objects.filter(budget=approved_budget)
            bac   = float(items.aggregate(t=Sum('base_cost'))['t'] or 0)

            # Current week relative to project start
            current_week = 9999
            if project.start_date:
                days_elapsed = (today - project.start_date).days
                current_week = max(1, (days_elapsed // 7) + 1)

            pv = float(items.filter(week_no__lte=current_week).aggregate(t=Sum('base_cost'))['t'] or 0)

            # By category planned
            for row in items.values('category').annotate(planned=Sum('base_cost')):
                by_category[row['category']] = {'planned': float(row['planned'] or 0), 'actual': 0}

        # ── Actual Cost from WeeklyProgress ───────────────────────────────────
        wp = WeeklyProgress.objects.filter(project=project).aggregate(
            total      = Sum('total_actual'),
            materials  = Sum('materials_actual'),
            fuel       = Sum('fuel_actual'),
            labour     = Sum('labour_actual'),
            casuals    = Sum('casuals_actual'),
        )
        ac = float(wp['total'] or 0)

        for key in ('materials', 'fuel', 'labour', 'casuals'):
            if key in by_category:
                by_category[key]['actual'] = float(wp[key] or 0)

        # ── Earned Value from IPC certified amounts ───────────────────────────
        ipc_agg = IPC.objects.filter(project=project).aggregate(
            claimed   = Sum('amount_claimed'),
            certified = Sum('amount_certified'),
            paid      = Sum('amount_paid'),
        )
        ev              = float(ipc_agg['certified'] or 0)
        revenue_claimed = float(ipc_agg['claimed']   or 0)
        revenue_paid    = float(ipc_agg['paid']       or 0)

        # ── Finance costs from finance module ─────────────────────────────────
        try:
            from finance.models import Bill, ExpenseClaim
            bills_total    = float(Bill.objects.filter(project=project).aggregate(t=Sum('total_amount'))['t'] or 0)
            expenses_total = float(ExpenseClaim.objects.filter(project=project, status='approved').aggregate(t=Sum('total_amount'))['t'] or 0)
        except Exception:
            bills_total    = 0.0
            expenses_total = 0.0

        # ── EVM metrics ───────────────────────────────────────────────────────
        cpi  = round(ev / ac,  3) if ac  > 0 else None
        spi  = round(ev / pv,  3) if pv  > 0 else None
        cv   = round(ev - ac,  2)
        sv   = round(ev - pv,  2)
        eac  = round(bac / cpi,  2) if cpi and cpi > 0 else round(bac, 2)
        vac  = round(bac - eac,  2)
        tcpi_denom = bac - ac
        tcpi = round((bac - ev) / tcpi_denom, 3) if tcpi_denom > 0 else None
        pct_complete = round(ev / bac * 100, 1) if bac > 0 else 0.0

        # ── Variation orders ──────────────────────────────────────────────────
        vos        = project.variation_orders.all()
        vo_total   = float(vos.aggregate(t=Sum('amount'))['t'] or 0)
        vo_approved = float(vos.filter(status='approved').aggregate(t=Sum('amount'))['t'] or 0)
        revised_cv  = float(project.contract_value or 0) + vo_approved

        # ── Weekly actuals for S-curve ────────────────────────────────────────
        weekly_actuals = list(
            WeeklyProgress.objects.filter(project=project)
            .order_by('week_no')
            .values('week_no', 'total_actual', 'materials_actual', 'fuel_actual', 'labour_actual', 'casuals_actual')
        )

        return Response({
            'project': {
                'id':                    str(project.id),
                'name':                  project.name,
                'contract_value':        float(project.contract_value or 0),
                'revised_contract_value': revised_cv,
                'start_date':            str(project.start_date) if project.start_date else None,
                'end_date':              str(project.end_date)   if project.end_date   else None,
                'status':                project.status,
            },
            'evm': {
                'bac':          bac,
                'pv':           pv,
                'ev':           ev,
                'ac':           ac,
                'cv':           cv,
                'sv':           sv,
                'cpi':          cpi,
                'spi':          spi,
                'eac':          eac,
                'vac':          vac,
                'tcpi':         tcpi,
                'pct_complete': pct_complete,
            },
            'revenue': {
                'contract_value': float(project.contract_value or 0),
                'claimed':   revenue_claimed,
                'certified': ev,
                'paid':      revenue_paid,
                'ipc_count': IPC.objects.filter(project=project).count(),
            },
            'costs': {
                'bac':             bac,
                'ac_weekly':       ac,
                'bills_total':     bills_total,
                'expenses_total':  expenses_total,
                'by_category':     by_category,
            },
            'variation_orders': {
                'count':           vos.count(),
                'total_amount':    vo_total,
                'approved_amount': vo_approved,
            },
            'weekly_actuals': weekly_actuals,
        })


class PortfolioSummaryView(APIView):
    """
    GET /api/v1/projects/portfolio/
    Cross-project executive roll-up: counts, contract value, IPC totals, EVM averages, risk summary.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        projects = Project.objects.all()

        by_status = {}
        for status_val, _ in Project.STATUS_CHOICES:
            by_status[status_val] = projects.filter(status=status_val).count()

        total_contract_value = projects.aggregate(s=Sum('contract_value'))['s'] or 0

        active_projects = projects.filter(status='active')

        # IPC totals across all active projects
        ipc_qs = IPC.objects.filter(project__in=active_projects)
        ipc_agg = ipc_qs.aggregate(
            claimed=Sum('amount_claimed'),
            certified=Sum('amount_certified'),
            paid=Sum('amount_paid'),
        )

        # Risk summary
        risks = ProjectRisk.objects.filter(project__in=active_projects)
        risk_summary = risks.aggregate(
            open=Count('id', filter=Q(status='open')),
            high=Count('id', filter=Q(likelihood='high', impact='high')),
            critical=Count('id', filter=Q(status='open', impact='critical') | Q(status='open', likelihood='high', impact='high')),
        )

        # Per-project health cards (active only, lightweight)
        project_cards = []
        for p in active_projects.order_by('name'):
            # Latest weekly progress % complete
            latest_wp = WeeklyProgress.objects.filter(project=p).order_by('-week_ending').first()
            pct = float(latest_wp.percent_complete) if latest_wp and latest_wp.percent_complete else None

            # IPC certified = EV proxy
            ev = float(IPC.objects.filter(project=p).aggregate(s=Sum('amount_certified'))['s'] or 0)
            bac = float(BudgetLineItem.objects.filter(budget__project=p, budget__status='approved').aggregate(s=Sum('base_cost'))['s'] or 0)
            ac = float(WeeklyProgress.objects.filter(project=p).aggregate(s=Sum('total_actual'))['s'] or 0)
            cpi = round(ev / ac, 2) if ac > 0 else None
            spi = round(ev / bac, 2) if bac > 0 else None

            project_cards.append({
                'id': str(p.id),
                'code': p.code,
                'name': p.name,
                'status': p.status,
                'contract_value': float(p.contract_value or 0),
                'pct_complete': pct,
                'ev': ev,
                'ac': ac,
                'bac': bac,
                'cpi': cpi,
                'spi': spi,
            })

        return Response({
            'totals': {
                'total_projects': projects.count(),
                'active_projects': active_projects.count(),
                'total_contract_value': float(total_contract_value),
                'by_status': by_status,
            },
            'ipc': {
                'claimed':   float(ipc_agg['claimed'] or 0),
                'certified': float(ipc_agg['certified'] or 0),
                'paid':      float(ipc_agg['paid'] or 0),
            },
            'risks': risk_summary,
            'project_cards': project_cards,
        })


# ── New Phase-2 views ──────────────────────────────────────────────────────────

from .models import (
    ChainageSegment, SiteDiary, QATestRecord, NonConformance,
    RFIRecord, IncidentReport, Subcontractor, SubcontractorMilestone,
)
from .serializers import (
    ChainageSegmentSerializer, SiteDiarySerializer, QATestRecordSerializer,
    NonConformanceSerializer, RFIRecordSerializer, IncidentReportSerializer,
    SubcontractorSerializer, SubcontractorMilestoneSerializer,
)


class ChainageSegmentListCreate(generics.ListCreateAPIView):
    serializer_class = ChainageSegmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ChainageSegment.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class ChainageSegmentDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ChainageSegmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ChainageSegment.objects.filter(project_id=self.kwargs['project_pk'])


class SiteDiaryListCreate(generics.ListCreateAPIView):
    serializer_class = SiteDiarySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SiteDiary.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project, prepared_by=self.request.user)


class SiteDiaryDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SiteDiarySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SiteDiary.objects.filter(project_id=self.kwargs['project_pk'])


class QATestRecordListCreate(generics.ListCreateAPIView):
    serializer_class = QATestRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return QATestRecord.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class QATestRecordDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = QATestRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return QATestRecord.objects.filter(project_id=self.kwargs['project_pk'])


class NonConformanceListCreate(generics.ListCreateAPIView):
    serializer_class = NonConformanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return NonConformance.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project, raised_by=self.request.user)


class NonConformanceDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = NonConformanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return NonConformance.objects.filter(project_id=self.kwargs['project_pk'])


class RFIRecordListCreate(generics.ListCreateAPIView):
    serializer_class = RFIRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return RFIRecord.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project, raised_by=self.request.user)


class RFIRecordDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RFIRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return RFIRecord.objects.filter(project_id=self.kwargs['project_pk'])


class IncidentReportListCreate(generics.ListCreateAPIView):
    serializer_class = IncidentReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return IncidentReport.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project, reported_by=self.request.user)


class IncidentReportDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = IncidentReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return IncidentReport.objects.filter(project_id=self.kwargs['project_pk'])


class SubcontractorListCreate(generics.ListCreateAPIView):
    serializer_class = SubcontractorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Subcontractor.objects.filter(project_id=self.kwargs['project_pk'])

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs['project_pk'])
        serializer.save(project=project)


class SubcontractorDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SubcontractorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Subcontractor.objects.filter(project_id=self.kwargs['project_pk'])


class SubcontractorMilestoneListCreate(generics.ListCreateAPIView):
    serializer_class = SubcontractorMilestoneSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SubcontractorMilestone.objects.filter(subcontractor_id=self.kwargs['sub_pk'])


class SubcontractorMilestoneDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SubcontractorMilestoneSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SubcontractorMilestone.objects.filter(subcontractor_id=self.kwargs['sub_pk'])
