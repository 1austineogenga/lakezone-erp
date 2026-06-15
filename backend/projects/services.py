import re
from decimal import Decimal, InvalidOperation

import openpyxl

from .models import BOQ, BOQBill, BOQItem, Budget, BudgetLineItem, IPC, Project, ProjectPersonnel, ProjectRisk


SKIP_KEYWORDS = [
    'carried forward', 'carry forward', 'brought forward', 'sub-total',
    'subtotal', 'total', 'sub total', 'page total', 'bill total',
    'grand total', 'summary',
]


def _looks_like_skip_row(description):
    if not description:
        return True
    desc_lower = str(description).strip().lower()
    if not desc_lower:
        return True
    for kw in SKIP_KEYWORDS:
        if kw in desc_lower:
            return True
    return False


def _safe_decimal(value, default=Decimal('0')):
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


class BOQImportService:
    @staticmethod
    def import_from_excel(file, project_id, title):
        project = Project.objects.get(pk=project_id)

        wb = openpyxl.load_workbook(file, data_only=True)

        boq = BOQ.objects.create(
            project=project,
            title=title,
        )

        bills_created = 0
        items_created = 0
        bill_order = 0

        # First pass: collect bill totals from summary sheet if present
        bill_totals = {}
        for sheet_name in wb.sheetnames:
            if 'summary' in sheet_name.lower():
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    # Look for rows that have a bill number in col A or B and amount in last cols
                    pass  # summary sheet totals are re-computed from items

        # Second pass: process individual bill sheets
        bill_pattern = re.compile(r'^bill\s*(\d+)$', re.IGNORECASE)

        for sheet_name in wb.sheetnames:
            match = bill_pattern.match(sheet_name.strip())
            if not match:
                continue

            bill_number = match.group(1)
            ws = wb[sheet_name]

            # Find the header row
            header_row_idx = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_vals = [str(c).strip().lower() if c is not None else '' for c in row]
                # Look for 'description' in B or C column (index 1 or 2)
                if len(row_vals) >= 3:
                    if 'description' in row_vals[1] or 'description' in row_vals[2]:
                        header_row_idx = row_idx
                        break

            if header_row_idx is None:
                # Try first few rows
                header_row_idx = 1

            # Get sheet title from first few rows
            bill_description = sheet_name
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell.strip()) > 5:
                        candidate = cell.strip()
                        if 'bill' in candidate.lower() or len(candidate) > 10:
                            bill_description = candidate
                            break
                else:
                    continue
                break

            bill = BOQBill.objects.create(
                boq=boq,
                bill_number=bill_number,
                description=bill_description[:500],
                order=bill_order,
            )
            bill_order += 1
            bills_created += 1

            bill_sub_total = Decimal('0')

            # Read items starting after header row
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if len(row) < 3:
                    continue

                item_no = row[0]
                description = row[1] if len(row) > 1 else None
                unit = row[2] if len(row) > 2 else None
                qty = row[3] if len(row) > 3 else None
                rate = row[4] if len(row) > 4 else None
                amount = row[5] if len(row) > 5 else None

                # Skip if no item number
                if not item_no or str(item_no).strip() == '':
                    continue

                item_no_str = str(item_no).strip()

                # Skip if description looks like a subtotal
                if _looks_like_skip_row(description):
                    continue

                # Skip if item_no looks like a label rather than a number
                if _looks_like_skip_row(item_no_str):
                    continue

                desc_str = str(description).strip() if description else ''
                if not desc_str:
                    continue

                qty_dec = _safe_decimal(qty)
                rate_dec = _safe_decimal(rate)
                amount_dec = _safe_decimal(amount)

                # If amount is zero but qty and rate are set, compute it
                if amount_dec == 0 and qty_dec != 0 and rate_dec != 0:
                    amount_dec = qty_dec * rate_dec

                BOQItem.objects.create(
                    bill=bill,
                    item_number=item_no_str[:20],
                    description=desc_str,
                    unit=str(unit).strip()[:50] if unit else '',
                    quantity=qty_dec,
                    rate=rate_dec,
                    amount=amount_dec,
                )
                items_created += 1
                bill_sub_total += amount_dec

            # Update bill sub_total
            bill.sub_total = bill_sub_total
            bill.save()

        total_amount = sum(
            bill.sub_total for bill in boq.bills.all()
        )

        return {
            'boq_id': str(boq.id),
            'bills_created': bills_created,
            'items_created': items_created,
            'total_amount': float(total_amount),
        }


import datetime

# ── Known project defaults per workbook prefix ─────────────────────────────────

_PROJECT_DEFAULTS = {
    'MN': {
        'name':            'Magumu - Njambini Road Rehabilitation',
        'client':          'Kenya National Highways Authority (KeNHA)',
        'contract_number': 'KeNHA/MN/2025',
        'contract_value':  Decimal('0'),        # unknown — fill manually
        'location':        'Magumu - Njambini, Nyandarua County',
        'status':          'active',
        'start_date':      datetime.date(2026, 2, 1),
        'end_date':        datetime.date(2026, 9, 30),
        'description':     (
            'Rehabilitation of Magumu - Njambini Road. '
            '2-month execution budget covers drainage (RC pipes, culverts), '
            'footpath paving, stone pitching and shoulder reconstruction.'
        ),
    },
    'NS': {
        'name':            'Njambini - Sasumua Dam Road Rehabilitation',
        'client':          'Kenya National Highways Authority (KeNHA)',
        'contract_number': 'KeNHA/NS/2025',
        # BOQ Sub-Total (2): A+B+C = KES 532,833,705.60 (excl. VAT)
        'contract_value':  Decimal('532833705.60'),
        'location':        'Njambini - Sasumua Dam Road, Nyandarua County',
        'status':          'active',
        'start_date':      datetime.date(2026, 2, 1),
        'end_date':        datetime.date(2027, 1, 31),
        'description':     (
            'Rehabilitation of Njambini - Sasumua Dam Road (4.71 km). '
            'Works include earthworks, cement-improved subbase/base, AC Type II asphalt, '
            'drainage culverts (Ø600/Ø900mm), concrete headwalls and road furniture. '
            'Contract value KES 532.8M excl. VAT (BOQ approved).'
        ),
    },
}

_NS_CATEGORY_MAP = {
    'earthworks / fill':           'materials',
    'pavement':                    'materials',
    'surfacing':                   'materials',
    'drainage':                    'materials',
    'operations':                  'other',
    'plant/vehicle pool':          'fuel',
    'small plant':                 'fuel',
    'plant support':               'other',
    'key personnel':               'labour',
    'casual labour':               'casuals',
    'operators/drivers support':   'labour',
}


class BudgetWorkbookImportService:
    """
    Parses a Combined Budget workbook (MN + NS format) and creates/updates
    Projects, Budgets, and BudgetLineItems.
    """

    @staticmethod
    def import_from_excel(file):
        wb = openpyxl.load_workbook(file, data_only=True)

        # Read global inputs
        budget_weeks = 8
        if 'Global_Inputs' in wb.sheetnames:
            ws_gi = wb['Global_Inputs']
            for row in ws_gi.iter_rows(min_row=4, values_only=True):
                if row[0] and 'budget period' in str(row[0]).lower():
                    try:
                        budget_weeks = int(row[1])
                    except (TypeError, ValueError):
                        pass

        # Detect project prefixes from sheet names (e.g. MN_Materials → MN)
        prefixes = set()
        for sheet_name in wb.sheetnames:
            for suffix in ('_Materials', '_Fuel', '_FuelPlant', '_Casuals', '_HR_Labour'):
                if sheet_name.endswith(suffix):
                    prefix = sheet_name[: len(sheet_name) - len(suffix)]
                    if prefix:
                        prefixes.add(prefix)

        results = []

        for prefix in sorted(prefixes):
            defaults = _PROJECT_DEFAULTS.get(prefix, {'name': f'{prefix} Project', 'client': ''})

            project, created = Project.objects.update_or_create(
                code=prefix,
                defaults={k: v for k, v in defaults.items()},
            )

            budget = Budget.objects.create(
                project=project,
                title=f'{defaults.get("name", prefix)} — 2-Month Execution Budget (Apr–May 2026)',
                period_weeks=budget_weeks,
                status='approved',
            )

            items_created = 0

            # Materials (MN weekly or NS flat)
            mat = f'{prefix}_Materials'
            if mat in wb.sheetnames:
                items_created += BudgetWorkbookImportService._parse_mn_materials(wb[mat], budget)

            # Fuel (MN weekly)
            fuel = f'{prefix}_Fuel'
            if fuel in wb.sheetnames:
                items_created += BudgetWorkbookImportService._parse_mn_fuel(wb[fuel], budget)

            # FuelPlant (NS flat)
            fuelplant = f'{prefix}_FuelPlant'
            if fuelplant in wb.sheetnames:
                items_created += BudgetWorkbookImportService._parse_ns_flat(wb[fuelplant], budget)

            # Casuals (MN weekly)
            casuals = f'{prefix}_Casuals'
            if casuals in wb.sheetnames:
                items_created += BudgetWorkbookImportService._parse_mn_casuals(wb[casuals], budget)

            # HR/Labour (NS flat)
            hr = f'{prefix}_HR_Labour'
            if hr in wb.sheetnames:
                items_created += BudgetWorkbookImportService._parse_ns_flat(wb[hr], budget)

            results.append({
                'project_id':   str(project.id),
                'project_code': prefix,
                'project_name': project.name,
                'created':      created,
                'budget_id':    str(budget.id),
                'items_created': items_created,
            })

        # ── Risks from Variance_Register ──────────────────────────────────────
        if 'Variance_Register' in wb.sheetnames:
            # Build lookup: code → project object
            project_map = {r['project_code']: Project.objects.get(pk=r['project_id']) for r in results}
            BudgetWorkbookImportService._parse_variance_register(
                wb['Variance_Register'], project_map
            )

        # ── Personnel from HR_Inputs_{PREFIX} ─────────────────────────────────
        for r in results:
            prefix = r['project_code']
            hr_sheet = f'HR_Inputs_{prefix}'
            if hr_sheet in wb.sheetnames:
                project = Project.objects.get(pk=r['project_id'])
                BudgetWorkbookImportService._parse_hr_inputs(wb[hr_sheet], project)

        return {'projects': results}

    # ── Sheet parsers ────────────────────────────────────────────────────────

    @staticmethod
    def _parse_mn_materials(ws, budget):
        """
        MN_Materials: col0=Week, col1=Month, col2=WorkFocus, col3=Material,
        col4=Qty, col5=Unit, col6=BaseRate, col7=Waste%, col8=LowVar%,
        col9=HighVar%, col10=BaseCost, col11=LowCase, col12=HighCase,
        col13=VarianceReserve, col14=Remarks
        """
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            week_no = row[0]
            if not isinstance(week_no, (int, float)):
                continue
            description = str(row[3]).strip() if row[3] else ''
            if not description:
                continue
            BudgetLineItem.objects.create(
                budget=budget,
                week_no=int(week_no),
                month_no=BudgetWorkbookImportService._month_no(row[1]),
                work_focus=(str(row[2]) if row[2] else '')[:200],
                category='materials',
                description=description[:500],
                quantity=_safe_decimal(row[4]),
                unit=(str(row[5]) if row[5] else '')[:50],
                base_rate=_safe_decimal(row[6]),
                waste_allowance_pct=_safe_decimal(row[7]),
                low_variance_pct=_safe_decimal(row[8]),
                high_variance_pct=_safe_decimal(row[9]),
                base_cost=_safe_decimal(row[10]),
                low_case_cost=_safe_decimal(row[11]),
                high_case_cost=_safe_decimal(row[12]),
                variance_reserve=_safe_decimal(row[13]),
                remarks=(str(row[14]) if row[14] else '')[:500],
            )
            count += 1
        return count

    @staticmethod
    def _parse_mn_fuel(ws, budget):
        """
        MN_Fuel: col0=Week, col1=Month, col2=WorkFocus, col3=Allocation,
        col4=FuelType, col5=Qty, col6=Unit, col7=BaseRate, col8=Allowance%,
        col9=LowVar%, col10=HighVar%, col11=BaseCost, col12=LowCase,
        col13=HighCase, col14=VarianceReserve, col15=Remarks
        """
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            week_no = row[0]
            if not isinstance(week_no, (int, float)):
                continue
            fuel_type = str(row[4]).strip() if row[4] else ''
            allocation = str(row[3]).strip() if row[3] else ''
            description = f'{allocation} — {fuel_type}' if allocation else fuel_type
            if not description.strip():
                continue
            BudgetLineItem.objects.create(
                budget=budget,
                week_no=int(week_no),
                month_no=BudgetWorkbookImportService._month_no(row[1]),
                work_focus=(str(row[2]) if row[2] else '')[:200],
                category='fuel',
                description=description[:500],
                quantity=_safe_decimal(row[5]),
                unit=(str(row[6]) if row[6] else '')[:50],
                base_rate=_safe_decimal(row[7]),
                waste_allowance_pct=_safe_decimal(row[8]),
                low_variance_pct=_safe_decimal(row[9]),
                high_variance_pct=_safe_decimal(row[10]),
                base_cost=_safe_decimal(row[11]),
                low_case_cost=_safe_decimal(row[12]),
                high_case_cost=_safe_decimal(row[13]),
                variance_reserve=_safe_decimal(row[14]),
                remarks=(str(row[15]) if len(row) > 15 and row[15] else '')[:500],
            )
            count += 1
        return count

    @staticmethod
    def _parse_mn_casuals(ws, budget):
        """
        MN_Casuals: col0=Week, col1=Month, col2=WorkFocus,
        col11=CasualHeadcount, col12=Days/Week, col13=DailyRate,
        col14=LowVar%, col15=HighVar%, col16=BaseCost,
        col17=LowCase, col18=HighCase, col19=Remarks
        """
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            week_no = row[0]
            if not isinstance(week_no, (int, float)):
                continue
            headcount = _safe_decimal(row[11])
            days_week = _safe_decimal(row[12])
            if headcount == 0:
                continue
            BudgetLineItem.objects.create(
                budget=budget,
                week_no=int(week_no),
                month_no=BudgetWorkbookImportService._month_no(row[1]),
                work_focus=(str(row[2]) if row[2] else '')[:200],
                category='casuals',
                description='Casual Labour',
                quantity=headcount * days_week,
                unit='person-days',
                base_rate=_safe_decimal(row[13]),
                low_variance_pct=_safe_decimal(row[14]),
                high_variance_pct=_safe_decimal(row[15]),
                base_cost=_safe_decimal(row[16]),
                low_case_cost=_safe_decimal(row[17]),
                high_case_cost=_safe_decimal(row[18]),
                remarks=(str(row[19]) if len(row) > 19 and row[19] else '')[:500],
            )
            count += 1
        return count

    @staticmethod
    def _parse_ns_flat(ws, budget):
        """
        NS flat sheets (NS_Materials, NS_FuelPlant, NS_HR_Labour):
        col0=Category, col1=Item, col2=Qty, col3=Unit, col4=BaseRate,
        col5=Variance%, col6=Notes, col7=BaseCost, col8=LowCase, col9=HighCase
        """
        count = 0
        current_category = 'other'
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not any(c for c in (row[:10] if len(row) >= 10 else row)):
                continue
            cat_cell = str(row[0]).strip() if row[0] else ''
            if cat_cell.upper() in ('TOTAL', 'SUB-TOTAL', 'SUBTOTAL', ''):
                if cat_cell:
                    continue  # skip total rows
            if cat_cell:
                current_category = _NS_CATEGORY_MAP.get(cat_cell.lower(), 'other')
            description = str(row[1]).strip() if row[1] else ''
            if not description:
                continue
            base_cost = _safe_decimal(row[7] if len(row) > 7 else None)
            qty = _safe_decimal(row[2] if len(row) > 2 else None)
            if base_cost == 0 and qty == 0:
                continue
            BudgetLineItem.objects.create(
                budget=budget,
                category=current_category,
                description=description[:500],
                quantity=qty,
                unit=(str(row[3]) if row[3] else '')[:50],
                base_rate=_safe_decimal(row[4] if len(row) > 4 else None),
                high_variance_pct=_safe_decimal(row[5] if len(row) > 5 else None),
                base_cost=base_cost,
                low_case_cost=_safe_decimal(row[8] if len(row) > 8 else None),
                high_case_cost=_safe_decimal(row[9] if len(row) > 9 else None),
                remarks=(str(row[6]) if row[6] else '')[:500],
            )
            count += 1
        return count

    @staticmethod
    def _parse_variance_register(ws, project_map):
        """
        Variance_Register sheet:
        col0=Project(MN/NS/Combined), col1=Risk, col2=ExpectedImpact,
        col3=BudgetTreatment, col4=RealisticRange, col5=Owner,
        col6=Status(Open/Pending), col7=Notes
        """
        _STATUS_MAP = {'open': 'open', 'pending': 'open', 'closed': 'closed', 'mitigated': 'mitigated', 'escalated': 'escalated'}
        _IMPACT_KEYWORDS = {
            'critical': ['critical', 'catastrophic'],
            'high': ['high', 'major', 'severe', 'significant'],
            'low': ['low', 'minor', 'negligible'],
        }

        def _guess_impact(risk_desc, impact_text):
            combined = (str(risk_desc) + ' ' + str(impact_text)).lower()
            for level in ('critical', 'high', 'low'):
                for kw in _IMPACT_KEYWORDS[level]:
                    if kw in combined:
                        return level
            return 'medium'

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            project_tag = str(row[0]).strip()
            if not project_tag or project_tag.lower() in ('project', 'combined*'):
                continue

            risk_desc = str(row[1]).strip() if row[1] else ''
            if not risk_desc:
                continue

            expected_impact = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            budget_treatment = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            realistic_range = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            owner = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            raw_status = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'open'
            notes = str(row[7]).strip() if len(row) > 7 and row[7] else ''

            status = _STATUS_MAP.get(raw_status, 'open')
            impact_level = _guess_impact(risk_desc, expected_impact)

            targets = []
            if project_tag.upper() == 'COMBINED':
                targets = list(project_map.values())
            elif project_tag.upper() in project_map:
                targets = [project_map[project_tag.upper()]]

            for project in targets:
                ProjectRisk.objects.update_or_create(
                    project=project,
                    risk_description=risk_desc,
                    defaults=dict(
                        expected_impact=expected_impact,
                        budget_treatment=budget_treatment,
                        realistic_range=realistic_range,
                        owner=owner,
                        impact_level=impact_level,
                        status=status,
                        notes=notes,
                    ),
                )

    _ROLE_MAP = {
        'site engineer': 'site_engineer',
        'site agent': 'site_engineer',
        'general foreman': 'general_foreman',
        'surveyor': 'surveyor',
        'foreman': 'foreman',
        'hse': 'hse_lead',
        'traffic': 'hse_lead',
        'clerk': 'clerk',
        'storekeeper': 'clerk',
        'timekeeper': 'clerk',
        'operator': 'operator',
        'driver': 'operator',
        'casual': 'casual',
    }

    @classmethod
    def _map_role(cls, role_text):
        lower = str(role_text).lower()
        for key, val in cls._ROLE_MAP.items():
            if key in lower:
                return val
        return 'other'

    @staticmethod
    def _parse_hr_inputs(ws, project):
        """
        HR_Inputs_{PREFIX}:
        col0=Project, col1=Category(Key personnel/Support/Allowance/Casual),
        col2=Role/LabourItem, col3=Headcount/Qty, col4=Months/DaysBasis,
        col5=RateKES, col6=Variance%, col7=BaseCostKES, col8=HighCaseKES,
        col9=Include?(Yes/No)
        """
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[2]:
                continue
            role_text = str(row[2]).strip()
            if not role_text or role_text.lower() in ('role / labour item', 'role', ''):
                continue
            headcount = _safe_decimal(row[3]) if len(row) > 3 else Decimal('1')
            if headcount == 0:
                headcount = Decimal('1')
            rate = _safe_decimal(row[5]) if len(row) > 5 else Decimal('0')
            include_raw = str(row[9]).strip().lower() if len(row) > 9 and row[9] else 'yes'
            include_in_budget = include_raw not in ('no', 'n', 'false', '0')
            base_cost = _safe_decimal(row[7]) if len(row) > 7 else Decimal('0')
            notes_parts = []
            cat = str(row[1]).strip() if row[1] else ''
            if cat:
                notes_parts.append(f'Category: {cat}')
            months = row[4]
            if months:
                notes_parts.append(f'Basis: {months}')
            variance = row[6]
            if variance:
                notes_parts.append(f'Variance: {variance}%')
            if base_cost:
                notes_parts.append(f'Base cost: KES {base_cost:,}')

            role = BudgetWorkbookImportService._map_role(role_text)

            ProjectPersonnel.objects.update_or_create(
                project=project,
                employee_name=role_text[:200],
                defaults=dict(
                    role=role,
                    monthly_rate=rate,
                    include_in_budget=include_in_budget,
                    notes='; '.join(notes_parts)[:500],
                    start_date=project.start_date,
                    end_date=project.end_date,
                ),
            )

    @staticmethod
    def _month_no(cell):
        if not cell:
            return None
        m = re.search(r'\d+', str(cell))
        return int(m.group()) if m else None
